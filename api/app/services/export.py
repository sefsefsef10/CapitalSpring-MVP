"""Export service for generating Excel reports."""

import io
import uuid
from datetime import datetime
from typing import Any, AsyncGenerator, Optional, Tuple

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import settings
from app.models.document import Document
from app.schemas.export import ExportTemplate
from app.services.storage import StorageService

logger = structlog.get_logger(__name__)


# Style definitions
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(size=10)
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12)

# Success/warning fills
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


class ExportService:
    """Service for generating Excel exports."""

    def __init__(self):
        """Initialize export service."""
        self.storage = StorageService()

    async def generate_excel(
        self,
        documents: list[Document],
        template: ExportTemplate,
        include_raw_data: bool = False,
        include_confidence_scores: bool = False,
        custom_fields: Optional[list[str]] = None,
    ) -> Tuple[str, int]:
        """
        Generate an Excel export from documents.

        Args:
            documents: List of documents to export
            template: Export template to use
            include_raw_data: Include raw extraction data
            include_confidence_scores: Include field confidence scores
            custom_fields: Custom field selection

        Returns:
            Tuple of (GCS path, file size in bytes)
        """
        logger.info(
            "Generating Excel export",
            template=template.value,
            document_count=len(documents),
        )

        # Create workbook
        wb = Workbook()

        # Generate content based on template
        if template == ExportTemplate.PORTFOLIO_FINANCIALS:
            self._generate_financials_export(wb, documents, include_confidence_scores)
        elif template == ExportTemplate.COVENANT_COMPLIANCE:
            self._generate_covenant_export(wb, documents, include_confidence_scores)
        elif template == ExportTemplate.BORROWING_BASE:
            self._generate_bbc_export(wb, documents, include_confidence_scores)
        elif template == ExportTemplate.CAPITAL_ACTIVITY:
            self._generate_capital_export(wb, documents, include_confidence_scores)
        elif template == ExportTemplate.EXCEPTION_REPORT:
            self._generate_exception_export(wb, documents)
        else:
            self._generate_custom_export(wb, documents, custom_fields, include_confidence_scores)

        # Include raw data sheet if requested
        if include_raw_data:
            self._add_raw_data_sheet(wb, documents)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        content = output.read()

        # Upload to GCS
        export_id = str(uuid.uuid4())
        gcs_path = f"gs://{settings.gcs_bucket_name}/exports/{export_id}.xlsx"

        # Upload using storage service
        blob_path = f"exports/{export_id}.xlsx"
        blob = self.storage.bucket.blob(blob_path)
        blob.upload_from_string(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        logger.info(
            "Excel export generated",
            gcs_path=gcs_path,
            size_bytes=len(content),
        )

        return gcs_path, len(content)

    def _generate_financials_export(
        self,
        wb: Workbook,
        documents: list[Document],
        include_confidence: bool = False,
    ) -> None:
        """Generate portfolio financials export."""
        ws = wb.active
        ws.title = "Financial Summary"

        # Title
        ws["A1"] = "Portfolio Company Financial Summary"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:H1")

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = SUBTITLE_FONT

        # Headers
        headers = [
            "Company", "Period", "Revenue", "Gross Profit", "Gross Margin",
            "EBITDA", "EBITDA Margin", "Net Income", "Total Assets", "Total Debt"
        ]
        if include_confidence:
            headers.append("Confidence")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER

        # Data rows
        for row, doc in enumerate(documents, 5):
            data = doc.extracted_data or {}

            values = [
                data.get("company_name", doc.original_filename),
                data.get("period_end_date", ""),
                self._format_currency(data.get("revenue")),
                self._format_currency(data.get("gross_profit")),
                self._format_percentage(data.get("gross_margin")),
                self._format_currency(data.get("ebitda")),
                self._format_percentage(data.get("ebitda_margin")),
                self._format_currency(data.get("net_income")),
                self._format_currency(data.get("total_assets")),
                self._format_currency(data.get("total_debt")),
            ]
            if include_confidence:
                values.append(self._format_percentage(doc.confidence * 100 if doc.confidence else 0))

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = DATA_FONT
                cell.border = BORDER
                if col >= 3:  # Numbers
                    cell.alignment = NUMBER_ALIGNMENT

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

    def _generate_covenant_export(
        self,
        wb: Workbook,
        documents: list[Document],
        include_confidence: bool = False,
    ) -> None:
        """Generate covenant compliance export."""
        ws = wb.active
        ws.title = "Covenant Compliance"

        # Title
        ws["A1"] = "Covenant Compliance Report"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:H1")

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Headers
        headers = [
            "Company", "Period", "Leverage Ratio", "Leverage Limit", "Leverage OK",
            "Interest Coverage", "Coverage Limit", "Coverage OK", "Overall Status"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER

        # Data rows
        for row, doc in enumerate(documents, 5):
            data = doc.extracted_data or {}

            leverage_ok = data.get("leverage_compliant", False)
            coverage_ok = data.get("coverage_compliant", False)
            overall_ok = data.get("overall_compliance", False)

            values = [
                data.get("company_name", doc.original_filename),
                data.get("reporting_period", ""),
                self._format_ratio(data.get("leverage_ratio")),
                self._format_ratio(data.get("leverage_covenant")),
                "PASS" if leverage_ok else "FAIL",
                self._format_ratio(data.get("interest_coverage_ratio")),
                self._format_ratio(data.get("coverage_covenant")),
                "PASS" if coverage_ok else "FAIL",
                "COMPLIANT" if overall_ok else "BREACH",
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = DATA_FONT
                cell.border = BORDER

                # Color-code compliance status
                if value == "PASS" or value == "COMPLIANT":
                    cell.fill = SUCCESS_FILL
                elif value == "FAIL" or value == "BREACH":
                    cell.fill = ERROR_FILL

        self._adjust_column_widths(ws)

    def _generate_bbc_export(
        self,
        wb: Workbook,
        documents: list[Document],
        include_confidence: bool = False,
    ) -> None:
        """Generate borrowing base certificate export."""
        ws = wb.active
        ws.title = "Borrowing Base"

        # Title
        ws["A1"] = "Borrowing Base Certificate Summary"
        ws["A1"].font = TITLE_FONT

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Headers
        headers = [
            "Company", "Date", "Gross AR", "Eligible AR", "AR Advance Rate",
            "Gross Inventory", "Eligible Inventory", "Inv Advance Rate",
            "Total Availability", "Outstanding", "Excess Availability"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER

        # Data rows
        for row, doc in enumerate(documents, 5):
            data = doc.extracted_data or {}

            values = [
                data.get("company_name", doc.original_filename),
                data.get("certificate_date", ""),
                self._format_currency(data.get("gross_accounts_receivable")),
                self._format_currency(data.get("eligible_ar")),
                self._format_percentage(data.get("ar_advance_rate")),
                self._format_currency(data.get("gross_inventory")),
                self._format_currency(data.get("eligible_inventory")),
                self._format_percentage(data.get("inventory_advance_rate")),
                self._format_currency(data.get("total_availability")),
                self._format_currency(data.get("outstanding_loans")),
                self._format_currency(data.get("excess_availability")),
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = DATA_FONT
                cell.border = BORDER

        self._adjust_column_widths(ws)

    def _generate_capital_export(
        self,
        wb: Workbook,
        documents: list[Document],
        include_confidence: bool = False,
    ) -> None:
        """Generate capital activity export."""
        ws = wb.active
        ws.title = "Capital Activity"

        # Title
        ws["A1"] = "Capital Activity Report"
        ws["A1"].font = TITLE_FONT

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Headers
        headers = [
            "Notice Date", "Due Date", "Call #", "Call Amount",
            "Purpose", "Cumulative Called", "Remaining Commitment"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER

        # Data rows
        for row, doc in enumerate(documents, 5):
            data = doc.extracted_data or {}

            values = [
                data.get("notice_date", ""),
                data.get("due_date", ""),
                data.get("call_number", ""),
                self._format_currency(data.get("call_amount")),
                data.get("call_purpose", ""),
                self._format_currency(data.get("cumulative_called")),
                self._format_currency(data.get("remaining_commitment")),
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = DATA_FONT
                cell.border = BORDER

        self._adjust_column_widths(ws)

    def _generate_exception_export(
        self,
        wb: Workbook,
        documents: list[Document],
    ) -> None:
        """Generate exception report export."""
        ws = wb.active
        ws.title = "Exceptions"

        # Title
        ws["A1"] = "Exception Report"
        ws["A1"].font = TITLE_FONT

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Headers
        headers = [
            "Document", "Category", "Field", "Reason",
            "Expected", "Actual", "Priority", "Status", "Created"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER

        # Data rows - iterate through document exceptions
        row = 5
        for doc in documents:
            for exc in doc.exceptions:
                values = [
                    doc.original_filename,
                    exc.category,
                    exc.field_name or "",
                    exc.reason,
                    exc.expected_value or "",
                    exc.actual_value or "",
                    exc.priority,
                    exc.status,
                    exc.created_at.strftime("%Y-%m-%d %H:%M") if exc.created_at else "",
                ]

                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = DATA_FONT
                    cell.border = BORDER

                    # Color-code priority
                    if col == 7:  # Priority column
                        if value == "critical":
                            cell.fill = ERROR_FILL
                        elif value == "high":
                            cell.fill = WARNING_FILL

                row += 1

        self._adjust_column_widths(ws)

    def _generate_custom_export(
        self,
        wb: Workbook,
        documents: list[Document],
        fields: Optional[list[str]],
        include_confidence: bool = False,
    ) -> None:
        """Generate custom export with specified fields."""
        ws = wb.active
        ws.title = "Data Export"

        # Title
        ws["A1"] = "Document Data Export"
        ws["A1"].font = TITLE_FONT

        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Determine fields to include
        if fields:
            export_fields = fields
        else:
            # Collect all unique fields from documents
            all_fields = set()
            for doc in documents:
                if doc.extracted_data:
                    all_fields.update(doc.extracted_data.keys())
            export_fields = ["filename", "doc_type", "status"] + sorted(all_fields)

        # Headers
        for col, field in enumerate(export_fields, 1):
            cell = ws.cell(row=4, column=col, value=field)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = BORDER

        # Data rows
        for row, doc in enumerate(documents, 5):
            data = doc.extracted_data or {}

            for col, field in enumerate(export_fields, 1):
                if field == "filename":
                    value = doc.original_filename
                elif field == "doc_type":
                    value = doc.doc_type
                elif field == "status":
                    value = doc.status
                else:
                    value = data.get(field, "")

                cell = ws.cell(row=row, column=col, value=str(value) if value else "")
                cell.font = DATA_FONT
                cell.border = BORDER

        self._adjust_column_widths(ws)

    def _add_raw_data_sheet(self, wb: Workbook, documents: list[Document]) -> None:
        """Add a sheet with raw extracted data as JSON."""
        import json

        ws = wb.create_sheet("Raw Data")

        ws["A1"] = "Raw Extracted Data"
        ws["A1"].font = TITLE_FONT

        headers = ["Document ID", "Filename", "Type", "Raw Data"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER

        for row, doc in enumerate(documents, 4):
            ws.cell(row=row, column=1, value=str(doc.id))
            ws.cell(row=row, column=2, value=doc.original_filename)
            ws.cell(row=row, column=3, value=doc.doc_type)
            ws.cell(row=row, column=4, value=json.dumps(doc.extracted_data or {}, indent=2))

    def _adjust_column_widths(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _format_currency(self, value: Any) -> str:
        """Format a value as currency."""
        if value is None:
            return ""
        try:
            return f"${float(value):,.2f}"
        except (ValueError, TypeError):
            return str(value)

    def _format_percentage(self, value: Any) -> str:
        """Format a value as percentage."""
        if value is None:
            return ""
        try:
            return f"{float(value):.1f}%"
        except (ValueError, TypeError):
            return str(value)

    def _format_ratio(self, value: Any) -> str:
        """Format a value as a ratio."""
        if value is None:
            return ""
        try:
            return f"{float(value):.2f}x"
        except (ValueError, TypeError):
            return str(value)

    async def get_download_url(self, gcs_path: str, expiration_hours: int = 24) -> str:
        """Get a signed download URL for an export."""
        return await self.storage.get_signed_url(gcs_path, expiration_hours)

    async def get_file_stream(
        self,
        export_id: uuid.UUID,
    ) -> Tuple[AsyncGenerator[bytes, None], str, str]:
        """Get file stream for download."""
        gcs_path = f"gs://{settings.gcs_bucket_name}/exports/{export_id}.xlsx"

        content = await self.storage.download_file(gcs_path)

        async def stream():
            yield content

        return (
            stream(),
            f"export_{export_id}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
